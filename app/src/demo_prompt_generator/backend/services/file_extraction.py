"""Best-effort text extraction for user-uploaded context files.

Single entry point: `extract_text(filename, data, max_chars=30_000)`. Returns
`(text, was_truncated)`. Dispatch is keyed on the file extension; unsupported
extensions raise ValueError so the route can map to a 400.

Token budget: extracted text is always truncated to `max_chars` before
returning, and tabular formats (CSV, XLSX) only read the first
`MAX_TABULAR_ROWS` rows up-front to avoid loading multi-million-row CSVs
into memory just to throw most of it away. The raw byte cap is enforced by
the caller (the route), not here — this module trusts what it's handed.
"""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path

# Extensions we know how to parse. Anything else returns an explicit error
# from the route so the user sees a readable "unsupported file type" chip
# instead of a silent empty extraction.
SUPPORTED_EXTENSIONS = frozenset(
    {
        ".pdf",
        ".xlsx",
        ".csv",
        ".docx",
        ".md",
        ".txt",
        ".json",
        ".yaml",
        ".yml",
        ".html",
        ".xml",
        ".log",
    }
)

# Tabular formats: only read this many data rows up-front. CSVs in the wild
# are often 100k+ rows; we just need a representative slice for the LLM to
# infer schema / domain.
MAX_TABULAR_ROWS = 10


def extract_text(
    filename: str,
    data: bytes,
    *,
    max_chars: int = 30_000,
) -> tuple[str, bool]:
    """Extract plain text from a file's bytes.

    Returns (text, was_truncated). `was_truncated` covers both the row cap
    (CSV / XLSX) and the character cap. Raises ValueError for unsupported
    extensions OR parser failures — the caller can map any ValueError to
    one HTTP 400 with a user-readable detail.
    """
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"unsupported file type: {ext or '(no extension)'}")

    # Each parser may raise anything (pypdf → PdfReadError, openpyxl →
    # InvalidFileException, docx → PackageNotFoundError, …). Catch broadly
    # and re-raise as ValueError so the route has ONE error shape to map.
    try:
        if ext == ".pdf":
            text, row_truncated = _extract_pdf(data, max_chars)
        elif ext == ".xlsx":
            text, row_truncated = _extract_xlsx(data)
        elif ext == ".csv":
            text, row_truncated = _extract_csv(data)
        elif ext == ".docx":
            text, row_truncated = _extract_docx(data), False
        elif ext == ".json":
            text, row_truncated = _extract_json(data), False
        else:
            # Plain-text formats — decode with errors='replace' so a stray
            # non-UTF8 byte doesn't blow up the whole extraction. The LLM
            # tolerates the occasional `�` replacement character fine.
            text, row_truncated = data.decode("utf-8", errors="replace"), False
    except ValueError:
        # Already shaped as a parser error by a deeper helper — passthrough.
        raise
    except Exception as e:
        raise ValueError(f"could not parse {ext} file: {e}") from e

    if len(text) > max_chars:
        text = text[:max_chars] + "\n\n[... truncated ...]"
        char_truncated = True
    else:
        char_truncated = False

    return text, (row_truncated or char_truncated)


def _extract_pdf(data: bytes, max_chars: int) -> tuple[str, bool]:
    """pypdf is pure Python, no system deps. Per-page text extraction; we
    bail early once we've collected ~5× max_chars worth (the caller still
    enforces the hard cap, but it's wasteful to parse 200 pages of a slide
    deck if we only ship the first 30 KB). 5× headroom covers whitespace
    + repeating headers we'll likely keep, while still bailing on huge
    documents."""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    total = 0
    truncated = False
    ceiling = 5 * max_chars
    for page in reader.pages:
        try:
            page_text = page.extract_text() or ""
        except Exception:
            # Some pages (e.g. image-only scans) raise inside pypdf — skip
            # them rather than failing the whole upload.
            page_text = ""
        parts.append(page_text)
        total += len(page_text)
        if total > ceiling:
            truncated = True
            break
    return "\n\n".join(p for p in parts if p), truncated


def _extract_xlsx(data: bytes) -> tuple[str, bool]:
    """openpyxl read-only mode + first sheet only. We render the first
    MAX_TABULAR_ROWS rows as a markdown-ish table so the LLM gets the
    column names AND a value sample (both useful for inferring domain)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = wb.active
    if sheet is None:
        return "(empty workbook)", False

    rows: list[list[str]] = []
    truncated = False
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= MAX_TABULAR_ROWS + 1:  # +1 to keep header AND 10 data rows
            truncated = True
            break
        rows.append(["" if cell is None else str(cell) for cell in row])

    if not rows:
        return "(empty sheet)", False

    lines = [f"Sheet: {sheet.title}"]
    for r in rows:
        lines.append("\t".join(r))
    return "\n".join(lines), truncated


def _extract_csv(data: bytes) -> tuple[str, bool]:
    """stdlib csv. First MAX_TABULAR_ROWS data rows (plus header). Same
    rationale as xlsx: schema + a sample is what the LLM needs."""
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows: list[list[str]] = []
    truncated = False
    for i, row in enumerate(reader):
        if i >= MAX_TABULAR_ROWS + 1:
            truncated = True
            break
        rows.append(row)
    if not rows:
        return "(empty file)", False
    return "\n".join("\t".join(r) for r in rows), truncated


def _extract_docx(data: bytes) -> str:
    """python-docx: paragraphs joined with blank lines."""
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_json(data: bytes) -> str:
    """Re-emit pretty-printed so the LLM sees structure (indented keys are
    easier to reason about than a single 5KB line). Fall back to raw if
    the file isn't actually valid JSON."""
    try:
        parsed = json.loads(data.decode("utf-8", errors="replace"))
        return json.dumps(parsed, indent=2, default=str)
    except (ValueError, UnicodeDecodeError):
        return data.decode("utf-8", errors="replace")
